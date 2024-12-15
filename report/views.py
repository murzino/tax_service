from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from django.http import FileResponse
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import math
from io import BytesIO

# Возвращаем шаблон домашней страницы
def home_view(request):
    return render(request, 'upload.html')

# Обработка файла
def process_file(file):
    df = pd.read_excel(file)

    # Извлечение данных
    branches = df['Филиал']
    employees = df['Сотрудник']
    tax_base = df['Налоговая база']
    calculated_total = df['Налог']

    # Логика расчета
    calculated_by_formula = []
    deviations = []
    for base, total in zip(tax_base, calculated_total):
        if base <= 5000000:
            formula = base * 0.13
        else:
            formula = (5000000 * 0.13) + ((base - 5000000) * 0.15)
        if not math.isnan(formula):
            formula = round(formula)
        calculated_by_formula.append(formula)

        if isinstance(total, int):
            deviations.append(total - formula)
        else:
            deviations.append(0)

    # Создание словаря на основе которого создадим новый файл
    report_data = {
        'ФИЛИАЛ': branches,
        'Сотрудник': employees,
        'Налоговая база': tax_base,
        'Налог': calculated_total,
        '': calculated_by_formula,
        'Отклонения': deviations
    }
    
    # Создание таблицы
    report_df = pd.DataFrame(report_data)

    # Вставка пустой строки в начало DataFrame(Потребуется что бы создать шапку по шаблону)
    report_df.loc[0] = [None] * len(report_df.columns)
    report_df.index = report_df.index + 1
    report_df = report_df.sort_index()

    # Отделяем последнюю строку
    last_row = report_df.iloc[[-1]]  # последняя строка
    df_to_sort = report_df.iloc[1:-1]  # все строки, кроме последней

    # Сортируем все, кроме последней строки
    sorted_df = df_to_sort.sort_values(by='Отклонения', ascending=False)

    # Объединяем отсортированные строки с последней строкой
    report_df = pd.concat([sorted_df, last_row])

    # Сохранение в BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name='Отчет')
        workbook = writer.book
        sheet = workbook.active

        # Изменение ширины и высоты
        sheet.column_dimensions['A'].width = 40.29
        sheet.column_dimensions['B'].width = 34.29
        sheet.column_dimensions['C'].width = 13.43
        sheet.column_dimensions['D'].width = 11.86
        sheet.column_dimensions['E'].width = 16.57
        sheet.column_dimensions['F'].width = 12.43
        sheet.row_dimensions[1].height = 12.75
        sheet.row_dimensions[2].height = 25.50
    
        sheet.cell(row=2, column=4, value="Исчислено всего")
        sheet.cell(row=2, column=5, value="Исчислено всего по формуле")
        sheet.cell(row=1, column=4, value="Налог")

        #Стили ячеек
        for col_idx in range(3, len(deviations)):  
                cell = sheet.cell(row=col_idx, column=6)
                value = report_df.iloc[col_idx-2, 5]
                if not math.isnan(value):
                    if value == 0:
                        cell.fill = PatternFill(fill_type='solid', fgColor="00FF00") 
                    else:
                        cell.fill = PatternFill(fill_type='solid', fgColor="FF0000") 
                else:
                    sheet.merge_cells(f"C{col_idx}:F{col_idx}")
                    sheet.cell(row=col_idx, column=3, value="У данного сотрудника отсутствует НБ")

        #Стили ячеек
        for i in range(1,3):
            for col_idx, value in enumerate(report_data, start=1):  
                cell = sheet.cell(row=i, column=col_idx)              
                cell.font = Font(name="Arial", bold=True, size=10)  
                cell.fill = PatternFill(fill_type="solid", fgColor='CBE4E5')
                cell.alignment = Alignment(
                    horizontal="center",  
                    vertical="center",   
                    wrap_text=True        
                )

                border = Border(
                    left=Side(border_style="thin", color="000000"),   
                    right=Side(border_style="thin", color="000000"),  
                    top=Side(border_style="thin", color="000000"),   
                    bottom=Side(border_style="thin", color="000000")  
                )
                cell.border = border

        sheet.merge_cells("A1:A2") 
        sheet.merge_cells("B1:B2") 
        sheet.merge_cells("C1:C2")  
        sheet.merge_cells("D1:E1")  
        sheet.merge_cells("F1:F2")  

    # Перемотка буфера в начало
    output.seek(0)
    return output

# Получение и возвращение обработаного файла
def upload_file_view(request):
    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']

        # Создаем экземпляр FileSystemStorage
        fs = FileSystemStorage()

        # Сохраняем загруженный файл
        filename = fs.save(file.name, file)
        uploaded_file_path = fs.path(filename)

        try:
            # Открываем файл в контекстном менеджере
            with open(uploaded_file_path, 'rb') as uploaded_file:
                processed_file = process_file(uploaded_file)
        finally:
            # Удаляем временные файлы
            os.remove(uploaded_file_path)
        
        # Возвращаем файл пользователю
        response = FileResponse(processed_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="processed_file.xlsx"'
        return response

    return render(request, 'upload.html')